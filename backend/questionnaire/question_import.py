"""
questionnaire/question_import.py — Parse uploaded CSV/XLSX files that DEFINE
the questions of a custom questionnaire.

This is distinct from importer.py (which imports vendor *answers*).
Here an auditor uploads a file describing the *questions* they want to ask.

Expected columns (case-insensitive, order-independent):
  - question_id    (optional; auto-generated q1, q2... if missing)
  - question_text  (required)
  - question_type  (optional; defaults to "free_text")
  - is_required    (optional; defaults to true — accepts yes/no/true/false/1/0)

Supported question_type values: free_text, boolean, numeric, file_attachment, multiple_choice.
Unknown types fall back to free_text.
"""

import csv
import io

VALID_TYPES = {"free_text", "boolean", "numeric", "file_attachment", "multiple_choice"}

# Header aliases → canonical field name
_ALIASES = {
    "question_id":   "question_id",
    "id":            "question_id",
    "qid":           "question_id",
    "question_text": "question_text",
    "question":      "question_text",
    "text":          "question_text",
    "prompt":        "question_text",
    "question_type": "question_type",
    "type":          "question_type",
    "answer_type":   "question_type",
    "is_required":   "is_required",
    "required":      "is_required",
    "mandatory":     "is_required",
}

_TRUE = {"true", "yes", "y", "1", "required", "t"}
_FALSE = {"false", "no", "n", "0", "optional", "f", ""}


class QuestionImportError(Exception):
    """Raised when an uploaded question file can't be parsed."""


def _canon_header(raw: str) -> str | None:
    if raw is None:
        return None
    key = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    return _ALIASES.get(key)


def _coerce_required(value) -> bool:
    if value is None:
        return True
    s = str(value).strip().lower()
    if s in _FALSE and s not in _TRUE:
        return False
    return True  # default to required when true or ambiguous


def _normalize_type(value) -> str:
    if value is None:
        return "free_text"
    s = str(value).strip().lower().replace(" ", "_").replace("-", "_")
    # Map common variants onto the canonical types used by the DB enum + UI.
    ALIASES = {
        "file_upload": "file_attachment", "file": "file_attachment",
        "upload": "file_attachment", "attachment": "file_attachment",
        "single_select": "multiple_choice", "select": "multiple_choice",
        "choice": "multiple_choice", "multiple_choice": "multiple_choice",
        "number": "numeric", "int": "numeric", "integer": "numeric",
        "text": "free_text", "string": "free_text",
        "bool": "boolean", "yes_no": "boolean",
    }
    s = ALIASES.get(s, s)
    return s if s in VALID_TYPES else "free_text"


def _rows_to_questions(rows: list[dict]) -> list[dict]:
    """Turn canonicalized rows into question dicts, validating as we go."""
    questions = []
    seen_ids = set()
    for i, row in enumerate(rows):
        text = (row.get("question_text") or "").strip()
        if not text:
            continue  # skip blank lines silently

        qid = (row.get("question_id") or "").strip() or f"q{len(questions) + 1}"
        if qid in seen_ids:
            raise QuestionImportError(
                f"Duplicate question_id '{qid}' on row {i + 1}. IDs must be unique."
            )
        seen_ids.add(qid)

        questions.append({
            "question_id":   qid,
            "question_text": text,
            "question_type": _normalize_type(row.get("question_type")),
            "is_required":   _coerce_required(row.get("is_required")),
        })

    if not questions:
        raise QuestionImportError(
            "No questions found. Make sure your file has a 'question_text' column "
            "with at least one non-empty row."
        )
    return questions


def parse_questions_csv(content: str) -> list[dict]:
    """Parse a CSV file defining questionnaire questions."""
    try:
        reader = csv.reader(io.StringIO(content))
        raw_rows = [r for r in reader if any(str(c).strip() for c in r)]
    except Exception as e:
        raise QuestionImportError(f"Could not read CSV: {e}")

    if not raw_rows:
        raise QuestionImportError("The CSV file is empty.")

    header = [_canon_header(c) for c in raw_rows[0]]
    if "question_text" not in header:
        raise QuestionImportError(
            "CSV must include a 'question_text' (or 'question') column header."
        )

    rows = []
    for raw in raw_rows[1:]:
        row = {}
        for col_name, cell in zip(header, raw):
            if col_name:
                row[col_name] = cell
        rows.append(row)
    return _rows_to_questions(rows)


def parse_questions_xlsx(content: bytes) -> list[dict]:
    """Parse an XLSX file defining questionnaire questions (first sheet)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise QuestionImportError(
            "XLSX support is not installed on the server (openpyxl missing)."
        )
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise QuestionImportError(f"Could not open the spreadsheet: {e}")

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Drop fully-empty rows
    all_rows = [r for r in all_rows if any(c is not None and str(c).strip() for c in r)]
    if not all_rows:
        raise QuestionImportError("The spreadsheet is empty.")

    header = [_canon_header(c) for c in all_rows[0]]
    if "question_text" not in header:
        raise QuestionImportError(
            "Spreadsheet must include a 'question_text' (or 'question') column header "
            "in the first row."
        )

    rows = []
    for raw in all_rows[1:]:
        row = {}
        for col_name, cell in zip(header, raw):
            if col_name:
                row[col_name] = cell
        rows.append(row)
    return _rows_to_questions(rows)


def parse_questions_file(filename: str, content: bytes) -> list[dict]:
    """Dispatch on file extension. Returns a list of question dicts."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_questions_xlsx(content)
    if name.endswith(".csv") or name.endswith(".txt"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        return parse_questions_csv(text)
    # Try to sniff: CSV-ish if it decodes as text
    try:
        text = content.decode("utf-8-sig")
        return parse_questions_csv(text)
    except Exception:
        raise QuestionImportError(
            "Unsupported file type. Upload a .csv or .xlsx file."
        )


# ── Template generation (downloadable example) ────────────────────────────

TEMPLATE_HEADERS = ["question_id", "question_text", "question_type", "is_required"]

TEMPLATE_ROWS = [
    ["cc6.1", "Do you enforce MFA on all privileged accounts?", "boolean", "true"],
    ["cc6.2", "Is customer data encrypted at rest?", "boolean", "true"],
    ["cc6.3", "What is your patch management cadence?", "free_text", "true"],
    ["cc7.1", "Upload your most recent penetration test report.", "file_attachment", "false"],
    ["cc7.2", "Which SIEM platform do you use?", "multiple_choice", "false"],
]


def build_template_csv() -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerows(TEMPLATE_ROWS)
    return buf.getvalue()


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in TEMPLATE_ROWS:
        ws.append(row)

    widths = [14, 60, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
