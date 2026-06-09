"""
questionnaire/upload_parser.py — Parse uploaded CSV/XLSX into question definitions

Expected columns (case-insensitive, order-independent):
  question_id   — short id e.g. "cc6.1" (optional, auto-generated if blank)
  question_text — the question (required)
  question_type — boolean | free_text | numeric | file_attachment (default free_text)
  is_required   — yes/no/true/false (default yes)
"""

import csv
import io

VALID_TYPES = {"boolean", "free_text", "numeric", "file_attachment"}


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _coerce_required(val) -> bool:
    s = str(val).strip().lower()
    return s not in ("no", "false", "0", "n", "")


def _coerce_type(val) -> str:
    s = (str(val).strip().lower().replace(" ", "_")) if val else "free_text"
    return s if s in VALID_TYPES else "free_text"


def _rows_to_questions(rows: list[dict]) -> list[dict]:
    questions = []
    for i, row in enumerate(rows):
        # normalize keys
        r = { _norm_header(k): v for k, v in row.items() }
        text = (r.get("question_text") or r.get("question") or "").strip()
        if not text:
            continue
        questions.append({
            "question_id":   (r.get("question_id") or r.get("id") or f"q{i+1}").strip(),
            "question_text": text,
            "question_type": _coerce_type(r.get("question_type") or r.get("type")),
            "is_required":   _coerce_required(r.get("is_required") or r.get("required") or "yes"),
        })
    if not questions:
        raise ValueError("No valid questions found. Ensure a 'question_text' column exists.")
    return questions


def parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return _rows_to_questions(list(reader))


def parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty spreadsheet.")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    dict_rows = []
    for row in rows[1:]:
        dict_rows.append({ headers[i]: row[i] for i in range(len(headers)) if i < len(row) })
    return _rows_to_questions(dict_rows)


def parse_upload(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(content)
    if name.endswith(".csv"):
        return parse_csv(content)
    # Try CSV as a fallback
    try:
        return parse_csv(content)
    except Exception:
        raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")
